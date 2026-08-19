"""
Swiggy/Scootsy PO parser — handles BOTH attachment generations.

Source: Gmail label SWIGGY_PO; sender domains: scootsy.com, swiggy.in

Two formats are in circulation and both must keep working. Swiggy switched over around
2026-08-06 without notice; every PO from that date failed with "No .xls attachment
found" because the new file is .xlsx and the old detector matched only .xls.

  LEGACY  SOTY-{SELLER_CODE}-{PO_NUMBER}.xls
          Microsoft SpreadsheetML — XML with an .xls extension, NOT binary xls.
          Read as a flat ordered list of non-empty cells (positional).

  CURRENT {CODE}_CREATE_OTB_PURCHASE_ORDER_{uuid}.xlsx
          Genuine OOXML (a ZIP), unreadable by the XML path. Read with openpyxl as a
          real 2D grid, and addressed by *column header name* rather than position —
          see _ooxml_column_map(). Position-indexing is what made the legacy path
          brittle, and this format has merged cells with a two-row header, so names
          are both safer and clearer.

Dispatch is on the file's magic bytes, not its extension: a ZIP starts with 'PK\x03\x04'.
Extensions have already proved unreliable here, and content cannot lie.

XLS flat-cell layout after XML parse:
  [0]  'Purchase Order'
  [1]  'Vendor Name :'
  [2]  'PO No :'   [3] po_number   [4] 'PO Date :'   [5] date_str   ...
  [10] 'Expected Delivery Date:'  [11] date_str  ...
  [14] 'Reference PO Code:'   [15] vendor_address
  [16] 'Billing Address'
  [17] 'Shipping Address'     [18] ship_addr
  [20] 'S.'  [21] 'Item Code' ... [42] 'Amt (INR)'   ← 23 column-header cells
  [43] '1'   [44] sku  [45] desc  ... [60] total      ← 18 cells per line item
  ...
  (footer row: large decimal, not a line number)
  ... 'Grand Total (INR)'  grand_total_value

Known quirks:
  - All sampled POs use IGST (interstate); CGST+SGST columns are always 0.
  - Item descriptions have embedded \\n from merged cells (replaced with space).
  - Prices use 5-digit precision; we round to 6dp for Decimal safety.
  - Subject pattern: '{CITY} {CODE}-{PO_NUMBER}-{VENDOR_NAME}'.
  - Vendor address is at flat-cell [15] (after 'Reference PO Code:' label row).
"""
from __future__ import annotations

import html
import re
import uuid
import xml.etree.ElementTree as ET
from datetime import date
from decimal import ROUND_HALF_UP, Decimal
from typing import Any

import requests
import structlog

from app.models._enums import PoStatus, SourceChannel
from app.parsers.base import BaseParser, ParseResult
from app.schemas.canonical import EDI850, EDI850Line, EDIAddress

log = structlog.get_logger(__name__)

_SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"
_ZERO = Decimal("0")
_TWO_DP = Decimal("0.01")
_COLS_PER_LINE = 18  # values per line-item row in the flat cell list
_HEADER_COLS = 23   # 'S.' … 'Amt (INR)' (indices 20-42 in sample)


class SwiggyParser(BaseParser):
    """
    Parses Swiggy/Scootsy POs from SpreadsheetML .xls email attachments.
    Downloads the .xls from its Cloudinary URL stored in attachment_paths.
    """

    @property
    def partner_code(self) -> str:
        return "SWIGGY"

    def can_parse(self, raw_message: Any) -> bool:
        paths = raw_message.attachment_paths or []
        if not isinstance(paths, list):
            return False
        return _find_spreadsheet(paths) is not None

    def parse(self, raw_message: Any) -> ParseResult:
        try:
            return self._do_parse(raw_message)
        except Exception as exc:
            log.exception("swiggy_parser.error", raw_id=str(getattr(raw_message, "id", "")))
            return ParseResult(
                success=False,
                errors=[f"Unexpected parse error: {exc}"],
                parser_name="SwiggyParser",
            )

    # ── Internal ──────────────────────────────────────────────────────────────

    def _do_parse(self, raw_message: Any) -> ParseResult:
        paths = raw_message.attachment_paths or []
        xls_att = _find_spreadsheet(paths)
        if not xls_att:
            seen = [
                a.get("filename", "?") for a in paths if isinstance(a, dict)
            ] or ["(none)"]
            return ParseResult(
                success=False,
                errors=[
                    "No .xls/.xlsx attachment found — cannot parse Swiggy PO. "
                    f"Attachments on this email: {', '.join(seen)}"
                ],
                parser_name="SwiggyParser",
            )

        url: str = xls_att.get("url", "")
        if not url:
            return ParseResult(
                success=False,
                errors=["Attachment has no URL"],
                parser_name="SwiggyParser",
            )

        content = _download(url)
        if content is None:
            return ParseResult(
                success=False,
                errors=[f"Failed to download attachment from {url}"],
                parser_name="SwiggyParser",
            )

        # Dispatch on content, not extension — see the module docstring.
        if _is_ooxml(content):
            fmt = "xlsx"
            extracted, extract_errors = _extract_ooxml(content)
        else:
            fmt = "spreadsheetml"
            extracted, extract_errors = _extract_spreadsheetml(content)

        if extracted is None:
            return ParseResult(
                success=False,
                errors=[f"Could not read the {fmt} attachment", *extract_errors],
                parser_name="SwiggyParser",
            )

        po_number = extracted.po_number or _po_from_filename(xls_att.get("filename", ""))
        if not po_number:
            return ParseResult(
                success=False,
                errors=[f"Cannot determine PO number from the {fmt} file or its filename"],
                parser_name="SwiggyParser",
            )

        po_date = extracted.po_date
        delivery_date = extracted.delivery_date
        ship_addr_raw = extracted.ship_to_raw

        lines, line_errors = extracted.lines, list(extract_errors)
        if not lines:
            return ParseResult(
                success=False,
                errors=[f"No line items found in the {fmt} file"] + line_errors,
                parser_name="SwiggyParser",
            )

        subtotal = _sum_decimal(li.taxable_amount for li in lines if li.taxable_amount)
        cgst_total = _sum_decimal(li.cgst_amount for li in lines if li.cgst_amount)
        sgst_total = _sum_decimal(li.sgst_amount for li in lines if li.sgst_amount)
        igst_total = _sum_decimal(li.igst_amount for li in lines if li.igst_amount)
        grand_total = subtotal + cgst_total + sgst_total + igst_total

        # Prefer the file's own footer grand total — it is what the retailer will
        # reconcile against, including any rounding we would not reproduce by summing.
        if extracted.grand_total:
            grand_total = extracted.grand_total

        doc = EDI850(
            id=uuid.uuid4(),
            correlation_id=uuid.uuid4(),
            trading_partner_code="SWIGGY",
            source_channel=SourceChannel.EMAIL,
            raw_message_id=getattr(raw_message, "id", None),
            buyer_po_number=po_number,
            buyer_po_date=po_date,
            requested_delivery_date=delivery_date,
            ship_to=_parse_ship_to(ship_addr_raw),
            buyer_name="Scootsy Logistics Private Limited",
            subtotal_amount=subtotal or None,
            cgst_amount=cgst_total or None,
            sgst_amount=sgst_total or None,
            igst_amount=igst_total or None,
            grand_total=grand_total or None,
            line_items=lines,
            po_status=PoStatus.PARSED,
        )

        log.info(
            "swiggy_parser.success",
            po_number=po_number,
            line_count=len(lines),
            grand_total=str(grand_total),
            source_format=fmt,
        )
        return ParseResult(
            success=True,
            doc=doc,
            warnings=line_errors,
            parser_name="SwiggyParser",
        )


# ── Attachment selection & format dispatch ────────────────────────────────────

_SPREADSHEET_EXTS = (".xlsx", ".xls")


def _find_spreadsheet(paths: list[Any]) -> dict[str, Any] | None:
    """
    Pick the spreadsheet attachment, preferring .xlsx.

    Both generations ship a .pdf alongside the data file, and the PDF is never the one
    we want. Order matters: ".xls" is a prefix of ".xlsx", so a naive endswith(".xls")
    check misses the current format entirely — which is exactly the bug that silently
    failed 183 POs from 2026-08-06 onward.
    """
    files = [a for a in paths if isinstance(a, dict) and a.get("url")]
    for ext in _SPREADSHEET_EXTS:
        for att in files:
            if att.get("filename", "").lower().endswith(ext):
                return att
    return None


def _is_ooxml(content: bytes) -> bool:
    """OOXML (.xlsx) is a ZIP; SpreadsheetML is plain XML text."""
    return content[:4] == b"PK\x03\x04"


class _Sheet:
    """Format-neutral extraction result, so assembly of the EDI850 stays in one place."""

    __slots__ = ("po_number", "po_date", "delivery_date", "ship_to_raw", "lines", "grand_total")

    def __init__(
        self,
        po_number: str | None = None,
        po_date: date | None = None,
        delivery_date: date | None = None,
        ship_to_raw: str | None = None,
        lines: list[EDI850Line] | None = None,
        grand_total: Decimal | None = None,
    ) -> None:
        self.po_number = po_number
        self.po_date = po_date
        self.delivery_date = delivery_date
        self.ship_to_raw = ship_to_raw
        self.lines = lines or []
        self.grand_total = grand_total


def _extract_spreadsheetml(content: bytes) -> tuple[_Sheet | None, list[str]]:
    """LEGACY path — flat positional cell list. Behaviour unchanged."""
    cells = _flat_cells(content)
    if not cells:
        return None, ["SpreadsheetML XML was empty or invalid"]

    lines, errors = _parse_line_items(cells)
    return _Sheet(
        po_number=_extract_after(cells, "PO No :"),
        po_date=_parse_date_flexible(_extract_after(cells, "PO Date :")),
        delivery_date=_parse_date_flexible(_extract_after(cells, "Expected Delivery Date:")),
        ship_to_raw=_extract_after(cells, "Shipping Address"),
        lines=lines,
        grand_total=_extract_footer_grand_total(cells),
    ), errors


# ── Current format: OOXML .xlsx ───────────────────────────────────────────────

def _extract_ooxml(content: bytes) -> tuple[_Sheet | None, list[str]]:
    """
    CURRENT path — real .xlsx read as a 2D grid.

    openpyxl is loaded WITHOUT read_only: Swiggy's generator writes an inaccurate
    <dimension> record, and read_only trusts it, reporting a 1x1 sheet for a file that
    actually holds 87 rows.
    """
    import io

    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001 - surfaced to the caller as a parse failure
        return None, [f"openpyxl could not open the .xlsx: {exc}"]

    ws = wb[wb.sheetnames[0]]
    rows: list[tuple[Any, ...]] = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, ["The .xlsx contained no rows"]

    header_idx = _ooxml_header_row(rows)
    if header_idx is None:
        return None, ["Could not locate the 'Item Code' header row in the .xlsx"]

    colmap = _ooxml_column_map(rows, header_idx)
    lines, errors = _ooxml_lines(rows, header_idx, colmap)

    return _Sheet(
        po_number=_ooxml_labelled(rows, "PO No"),
        po_date=_parse_date_flexible(_ooxml_labelled(rows, "PO Date")),
        delivery_date=_parse_date_flexible(_ooxml_labelled(rows, "Expected Delivery Date")),
        ship_to_raw=_ooxml_shipping_address(rows),
        lines=lines,
        grand_total=_ooxml_grand_total(rows),
    ), errors


def _cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _ooxml_labelled(rows: list[tuple[Any, ...]], label: str) -> str | None:
    """
    Value for a 'Label :' header field.

    Labels and values sit in the same row but in distant columns (label at ~11, value
    at ~18) with merged blanks between, so the value is the next non-empty cell to the
    right rather than a fixed offset. Trailing ':' and whitespace vary between files,
    hence the normalised comparison.
    """
    want = label.lower().rstrip(" :")
    for row in rows:
        for i, raw in enumerate(row):
            if _cell(raw).lower().rstrip(" :") != want:
                continue
            for nxt in row[i + 1:]:
                if _cell(nxt):
                    return _cell(nxt)
    return None


def _ooxml_shipping_address(rows: list[tuple[Any, ...]]) -> str | None:
    """
    Shipping address block, which sits in the row *below* its heading.

    'Billing Address' and 'Shipping Address' are side-by-side headings; the addresses
    are underneath in the matching columns. We take the cell directly below the
    'Shipping Address' heading, falling back to billing — an order still has to be
    parseable when only one address block is populated.
    """
    for r, row in enumerate(rows):
        for i, raw in enumerate(row):
            if _cell(raw).lower() != "shipping address":
                continue
            if r + 1 < len(rows):
                below = rows[r + 1]
                if i < len(below) and _cell(below[i]):
                    return _cell(below[i])
                for cell in below:
                    if _cell(cell):
                        return _cell(cell)
    return None


def _ooxml_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    """Index of the row containing the 'Item Code' column heading."""
    for r, row in enumerate(rows):
        if any(_cell(c).lower() == "item code" for c in row):
            return r
    return None


def _ooxml_column_map(rows: list[tuple[Any, ...]], header_idx: int) -> dict[str, int]:
    """
    Map logical field -> column index, by reading the header names.

    The header spans two rows: group labels ('CGST', 'SGST/UGST', 'IGST') on the first,
    sub-labels ('Rate', 'Amt (INR)') on the second, with merged cells leaving blanks.
    Carrying the last non-empty group forward reconstructs which group each sub-column
    belongs to, so 'CGST'+'Amt' resolves to a real column.

    Reading by name rather than fixed offsets is deliberate: this file already shifts
    columns between the item rows and the totals row, and Swiggy has changed the layout
    once without warning.
    """
    header = rows[header_idx]
    sub = rows[header_idx + 1] if header_idx + 1 < len(rows) else ()

    groups: list[str] = []
    current = ""
    for c in header:
        text = _cell(c)
        if text:
            current = text
        groups.append(current)

    colmap: dict[str, int] = {}

    def find(pred) -> int | None:
        for i in range(len(groups)):
            if pred(groups[i], _cell(sub[i]) if i < len(sub) else ""):
                return i
        return None

    simple = {
        "item_code": "item code",
        "description": "item desc",
        "hsn": "hsn code",
        "qty": "qty",
        "mrp": "mrp",
        "unit_price": "unit base cost",
        "taxable": "taxable value",
        "total": "total (inr)",
    }
    for key, name in simple.items():
        idx = find(lambda g, _s, n=name: g.lower().startswith(n))
        if idx is not None:
            colmap[key] = idx

    for key, group in (("cgst", "cgst"), ("sgst", "sgst"), ("igst", "igst")):
        rate = find(lambda g, s, gr=group: g.lower().startswith(gr) and s.lower() == "rate")
        amt = find(lambda g, s, gr=group: g.lower().startswith(gr) and s.lower().startswith("amt"))
        if rate is not None:
            colmap[f"{key}_rate"] = rate
        if amt is not None:
            colmap[f"{key}_amt"] = amt

    return colmap


def _ooxml_lines(
    rows: list[tuple[Any, ...]], header_idx: int, colmap: dict[str, int]
) -> tuple[list[EDI850Line], list[str]]:
    """
    Build line items from the grid.

    A data row is one whose first column is a plain integer serial number — that is
    what separates the ~39 item rows from the totals row beneath them, which leaves
    column 0 blank and would otherwise be read as a line with no SKU.
    """
    lines: list[EDI850Line] = []
    errors: list[str] = []

    def get(row: tuple[Any, ...], key: str) -> str:
        i = colmap.get(key)
        return _cell(row[i]) if i is not None and i < len(row) else ""

    for row in rows[header_idx + 1:]:
        serial = _cell(row[0] if row else "")
        if not serial.isdigit():
            continue

        item_code = get(row, "item_code")
        if not item_code:
            continue

        line_no = int(serial)
        try:
            qty = _to_decimal(get(row, "qty"))
            if qty <= _ZERO:
                errors.append(f"Line {line_no} ({item_code}): quantity is {qty}, skipped")
                continue

            taxable = _to_decimal(get(row, "taxable"))
            unit_price = _to_decimal(get(row, "unit_price"))
            if unit_price <= _ZERO and qty:
                unit_price = (taxable / qty).quantize(Decimal("0.000001"), ROUND_HALF_UP)

            cgst_amt = _to_decimal(get(row, "cgst_amt"))
            sgst_amt = _to_decimal(get(row, "sgst_amt"))
            igst_amt = _to_decimal(get(row, "igst_amt"))
            total = _to_decimal(get(row, "total")) or (taxable + cgst_amt + sgst_amt + igst_amt)

            lines.append(EDI850Line(
                line_number=line_no,
                buyer_sku=item_code,
                buyer_sku_description=get(row, "description").replace("\n", " ").strip() or None,
                hsn_code=get(row, "hsn") or None,
                ordered_qty=qty,
                buyer_uom="PC",
                unit_price=unit_price,
                # MRP is present in the file but EDI850Line has no field for it —
                # passing it would be silently dropped, so it is deliberately not sent.
                taxable_amount=taxable or None,
                cgst_rate=_to_decimal(get(row, "cgst_rate")) or None,
                cgst_amount=cgst_amt or None,
                sgst_rate=_to_decimal(get(row, "sgst_rate")) or None,
                sgst_amount=sgst_amt or None,
                igst_rate=_to_decimal(get(row, "igst_rate")) or None,
                igst_amount=igst_amt or None,
                line_total=total.quantize(_TWO_DP, ROUND_HALF_UP) if total else None,
            ))
        except Exception as exc:  # noqa: BLE001 - one bad row must not lose the PO
            errors.append(f"Line {line_no} ({item_code}): {exc}")

    return lines, errors


def _ooxml_grand_total(rows: list[tuple[Any, ...]]) -> Decimal | None:
    """
    Grand total from the labelled footer row.

    Taken from the 'Grand Total (INR)' label rather than the unlabelled totals row
    above it — that row's columns are offset from the item rows by a merged cell, so
    reading it positionally picks up the wrong figure.
    """
    for row in rows:
        for i, raw in enumerate(row):
            if _cell(raw).lower().startswith("grand total"):
                for nxt in row[i + 1:]:
                    value = _to_decimal(_cell(nxt))
                    if value > _ZERO:
                        return value
    return None


# ── XLS download & parse ──────────────────────────────────────────────────────

def _download(url: str) -> bytes | None:
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        return resp.content
    except Exception as exc:
        log.warning("swiggy_parser.download_failed", url=url, error=str(exc))
        return None


def _flat_cells(content: bytes) -> list[str]:
    """
    Parse SpreadsheetML XML and return a flat ordered list of non-empty cell values.
    Handles the Scootsy XLS which uses XML SpreadsheetML format with .xls extension.
    """
    try:
        text = content.decode("utf-8", errors="replace")
        root = ET.fromstring(text)
    except ET.ParseError as exc:
        log.warning("swiggy_parser.xml_parse_failed", error=str(exc))
        return []

    cells: list[str] = []
    for ws in root.iter(f"{{{_SS_NS}}}Worksheet"):
        for row in ws.iter(f"{{{_SS_NS}}}Row"):
            for cell in row.iter(f"{{{_SS_NS}}}Cell"):
                data = cell.find(f"{{{_SS_NS}}}Data")
                if data is not None and data.text:
                    val = html.unescape(data.text).replace("\n", " ").strip()
                    if val:
                        cells.append(val)
    return cells


# ── Cell extraction helpers ───────────────────────────────────────────────────

def _extract_after(cells: list[str], label: str) -> str | None:
    """Find a label in the flat cells list and return the immediately following cell."""
    label_norm = label.strip().lower()
    for i, c in enumerate(cells):
        if c.strip().lower() == label_norm and i + 1 < len(cells):
            return cells[i + 1]
    return None


def _po_from_filename(filename: str) -> str | None:
    """Extract PO number from 'SOTY-{SELLER_CODE}-{PO_NUMBER}.xls'."""
    stem = filename.rsplit(".", 1)[0]
    parts = stem.split("-")
    if len(parts) >= 3:
        return parts[-1]
    return None


def _parse_ship_to(raw: str | None) -> EDIAddress | None:
    if not raw:
        return None
    # Split first line as name, rest as address
    lines = [part.strip() for part in raw.split(",") if part.strip()]
    if not lines:
        return None
    name = "Scootsy Logistics Private Limited"
    return EDIAddress(
        name=name,
        line1=raw[:200] if raw else None,
    )


def _parse_date_flexible(value: str | None) -> date | None:
    if not value:
        return None
    # ISO datetime: '2026-07-25T00:00:00.000'
    if "T" in value:
        value = value.split("T")[0]
    # ISO date: '2026-07-25'
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        pass
    # English: 'Jul 14, 2026'
    import datetime
    for fmt in ("%b %d, %Y", "%d %b %Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue
    log.warning("swiggy_parser.unparsable_date", value=value)
    return None


# ── Line item parsing ─────────────────────────────────────────────────────────

def _parse_line_items(cells: list[str]) -> tuple[list[EDI850Line], list[str]]:
    """
    Find the column-header row ('S.' … 'Amt (INR)') then read 18-cell chunks.
    Stops when the first cell of a chunk is not a positive integer.
    """
    # Find 'S.' followed by 'Item Code' to locate the header row
    s_idx = -1
    for i, c in enumerate(cells):
        if c == "S." and i + 1 < len(cells) and cells[i + 1] == "Item Code":
            s_idx = i
            break
    if s_idx < 0:
        return [], ["Could not locate column headers in XLS"]

    items_start = s_idx + _HEADER_COLS  # skip 23 header cells
    lines: list[EDI850Line] = []
    errors: list[str] = []
    i = items_start

    while i + _COLS_PER_LINE <= len(cells):
        chunk = cells[i: i + _COLS_PER_LINE]
        # First cell is the line number (positive integer ≤ 999)
        try:
            line_no = int(float(chunk[0]))
            if line_no < 1 or line_no > 999:
                break
        except (ValueError, TypeError):
            break

        try:
            lines.append(_chunk_to_line(chunk, line_no))
        except Exception as exc:
            errors.append(f"Line {line_no} (sku={chunk[1] if len(chunk) > 1 else '?'}): {exc}")

        i += _COLS_PER_LINE

    return lines, errors


def _chunk_to_line(chunk: list[str], line_number: int) -> EDI850Line:
    """
    Map an 18-cell chunk to EDI850Line.

    Chunk layout (0-indexed):
      [0]  S.No          [1]  Item Code      [2]  Item Desc   [3] HSN
      [4]  Qty           [5]  MRP            [6]  Unit Cost   [7] Taxable Value
      [8]  CGST Rate     [9]  CGST Amt
      [10] SGST Rate     [11] SGST Amt
      [12] IGST Rate     [13] IGST Amt
      [14] CESS Rate     [15] CESS Amt
      [16] Add.CESS Amt  [17] Total
    """
    buyer_sku = chunk[1].strip()
    if not buyer_sku:
        raise ValueError("Empty item code")

    # Clean description: Scootsy wraps long names with embedded newlines already replaced
    desc_raw = chunk[2] if len(chunk) > 2 else ""
    description = re.sub(r"\s+", " ", desc_raw).strip() or None

    # Strip embedded product attributes noise: 'Colour:  Size: size Brand:CAMPAIGN'
    if description and "Colour:" in description:
        description = re.split(r"Colour:", description)[0].strip(" \t\n\r\x0b\x0c-")

    hsn = (chunk[3] if len(chunk) > 3 else "").strip() or None
    qty = _to_decimal(chunk[4] if len(chunk) > 4 else "0")
    if qty <= _ZERO:
        raise ValueError(f"Qty must be > 0, got {qty}")

    unit_price = _to_decimal(chunk[6] if len(chunk) > 6 else "0")
    taxable = _to_decimal(chunk[7] if len(chunk) > 7 else "0")

    cgst_rate = _to_decimal(chunk[8] if len(chunk) > 8 else "0")
    cgst_amt = _to_decimal(chunk[9] if len(chunk) > 9 else "0")
    sgst_rate = _to_decimal(chunk[10] if len(chunk) > 10 else "0")
    sgst_amt = _to_decimal(chunk[11] if len(chunk) > 11 else "0")
    igst_rate = _to_decimal(chunk[12] if len(chunk) > 12 else "0")
    igst_amt = _to_decimal(chunk[13] if len(chunk) > 13 else "0")
    cess_rate = _to_decimal(chunk[14] if len(chunk) > 14 else "0")
    cess_amt = _to_decimal(chunk[15] if len(chunk) > 15 else "0")
    line_total = _to_decimal(chunk[17] if len(chunk) > 17 else "0")

    return EDI850Line(
        line_number=line_number,
        buyer_sku=buyer_sku,
        buyer_sku_description=description,
        hsn_code=hsn,
        ordered_qty=qty,
        buyer_uom="EA",
        unit_price=unit_price if unit_price else None,
        taxable_amount=taxable if taxable else None,
        cgst_rate=cgst_rate if cgst_rate else None,
        cgst_amount=cgst_amt.quantize(_TWO_DP, ROUND_HALF_UP) if cgst_amt else None,
        sgst_rate=sgst_rate if sgst_rate else None,
        sgst_amount=sgst_amt.quantize(_TWO_DP, ROUND_HALF_UP) if sgst_amt else None,
        igst_rate=igst_rate if igst_rate else None,
        igst_amount=igst_amt.quantize(_TWO_DP, ROUND_HALF_UP) if igst_amt else None,
        cess_rate=cess_rate if cess_rate else None,
        cess_amount=cess_amt.quantize(_TWO_DP, ROUND_HALF_UP) if cess_amt else None,
        line_total=line_total.quantize(_TWO_DP, ROUND_HALF_UP) if line_total else None,
    )


def _extract_footer_grand_total(cells: list[str]) -> Decimal | None:
    """Find 'Grand Total (INR)' label and return the numeric value after it."""
    for i, c in enumerate(cells):
        if c.strip().lower() == "grand total (inr)" and i + 1 < len(cells):
            return _to_decimal(cells[i + 1]) or None
    return None


# ── Numeric helpers ───────────────────────────────────────────────────────────

def _to_decimal(value: Any) -> Decimal:
    if value is None:
        return _ZERO
    try:
        return Decimal(str(value).strip())
    except Exception:
        return _ZERO


def _sum_decimal(values: Any) -> Decimal:
    return sum((v for v in values if v is not None), _ZERO)
