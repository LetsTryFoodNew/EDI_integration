"""
Import real master data from the two ops-maintained Excel files, replacing
the dummy seed data.

Usage (inside the api container):
    python scripts/import_master_data.py <Mapping.xlsx> <sku_master.xlsx>

Source files (docs/):
  sku master.xlsx — one row per internal SKU:
    CATEGORY | SKU INTERNAL NAME | SKU SAP NAME | SAP ID | GRAMMAGE (g) |
    CASE SIZE | MRP (Rs.) | EAN | HSN | SHELF LIFE (Day) | Status | SKU IMAGE
  Mapping.xlsx — one row per (platform, platform SKU):
    CHAIN | ITEM CODE | ITEM NAME | SAP NAME | SAP ITEM CODE |
    SAP ALTERNATE CODE | MRP | Discount | GST RATE | UNIT COAST

Behaviour (idempotent — safe to re-run):
  - MaterialMaster upserted by b1_item_code; rows absent from the file are
    soft-deleted (dummy seed data disappears this way).
  - SkuMapping upserted by (partner, buyer_sku) as MANUALLY_MAPPED.
  - Mapping rows whose SAP ITEM CODE is '#N/A' get material_id=None and
    mapping_status=UNMAPPED so ops can resolve them in the dashboard.
"""
from __future__ import annotations

import sys
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import structlog
from sqlalchemy import select

from app.db import SyncSessionLocal
from app.models._enums import MappingStatus
from app.models.master_data import MaterialMaster, SkuMapping, TradingPartner

log = structlog.get_logger(__name__)

# Excel CHAIN value → trading_partners.code
_CHAIN_TO_PARTNER_CODE = {
    "SWIGGY": "SWIGGY",
    "ZEPTO": "ZEPTO",
    "FLIPKART": "FLIPKART",
    "BLINKIT": "BLINKIT",
    "BIGBASKET": "BIGBASKET",
    "AMAZON": "AMAZON",
    "RELIANCE": "RELIANCE_JIO",
}


def _clean(val: object) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def _load_sku_master(path: Path) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))[1:]
    items: list[dict[str, object]] = []
    for r in rows:
        sap_id = _clean(r[3])
        if not sap_id:
            continue
        items.append({
            "b1_item_code": sap_id.upper(),
            "description": _clean(r[2]) or _clean(r[1]) or sap_id,
            "grammage": r[4],
            "case_size": int(r[5]) if r[5] else None,
            "mrp": float(r[6]) if r[6] is not None else None,
            "ean": _clean(r[7]),
            "hsn_code": _clean(r[8]),
            "is_active": (_clean(r[10]) or "").upper() != "INACTIVE",
        })
    return items


def _load_mapping(path: Path) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))[1:]
    mappings: list[dict[str, object]] = []
    for r in rows:
        chain = _clean(r[0])
        buyer_sku = _clean(r[1])
        if not chain or not buyer_sku:
            continue
        sap_code = _clean(r[4])
        if sap_code and sap_code.upper() in ("#N/A", "N/A", "#REF!"):
            sap_code = None
        mappings.append({
            "chain": chain.upper(),
            "buyer_sku": buyer_sku,
            "buyer_sku_description": _clean(r[2]),
            "sap_item_code": sap_code.upper() if sap_code else None,
            "gst_rate": float(r[8]) if r[8] is not None else None,
        })
    return mappings


def import_materials(session, items: list[dict[str, object]]) -> tuple[int, int, int]:
    """Upsert MaterialMaster rows; soft-delete anything not in the file."""
    now = datetime.now(UTC)
    existing = {
        m.b1_item_code: m
        for m in session.execute(select(MaterialMaster)).scalars().all()
    }
    created = updated = 0
    file_codes = set()
    for item in items:
        code = item["b1_item_code"]
        file_codes.add(code)
        mat = existing.get(code)
        if mat is None:
            session.add(MaterialMaster(
                b1_item_code=code,
                description=item["description"],
                hsn_code=item["hsn_code"],
                uom="EA",
                case_size=item["case_size"],
                ean=item["ean"],
                mrp=item["mrp"],
                is_active=item["is_active"],
            ))
            created += 1
        else:
            mat.description = item["description"]
            mat.hsn_code = item["hsn_code"]
            mat.case_size = item["case_size"]
            mat.ean = item["ean"]
            mat.mrp = item["mrp"]
            mat.is_active = item["is_active"]
            mat.deleted_at = None
            updated += 1

    removed = 0
    for code, mat in existing.items():
        if code not in file_codes and mat.deleted_at is None:
            mat.deleted_at = now
            mat.is_active = False
            removed += 1
    session.flush()
    return created, updated, removed


def import_mappings(session, mappings: list[dict[str, object]]) -> tuple[int, int, int]:
    """Upsert SkuMapping rows keyed by (partner, buyer_sku)."""
    partners = {
        p.code: p
        for p in session.execute(select(TradingPartner)).scalars().all()
    }
    materials = {
        m.b1_item_code: m
        for m in session.execute(
            select(MaterialMaster).where(MaterialMaster.deleted_at.is_(None))
        ).scalars().all()
    }
    existing = {
        (sm.trading_partner_id, sm.buyer_sku): sm
        for sm in session.execute(select(SkuMapping)).scalars().all()
    }

    created = updated = unmapped = 0
    for row in mappings:
        partner_code = _CHAIN_TO_PARTNER_CODE.get(row["chain"])
        partner = partners.get(partner_code) if partner_code else None
        if partner is None:
            log.warning("import.unknown_chain", chain=row["chain"], buyer_sku=row["buyer_sku"])
            continue

        material = materials.get(row["sap_item_code"]) if row["sap_item_code"] else None
        if material is None:
            status = MappingStatus.UNMAPPED
            notes = (
                f"Imported from Mapping.xlsx — SAP code "
                f"{row['sap_item_code'] or 'missing (#N/A)'} not found in sku master"
            )
            unmapped += 1
        else:
            status = MappingStatus.MANUALLY_MAPPED
            notes = "Imported from Mapping.xlsx"
            if material.gst_rate is None and row["gst_rate"] is not None:
                material.gst_rate = row["gst_rate"]

        key = (partner.id, row["buyer_sku"])
        sm = existing.get(key)
        if sm is None:
            session.add(SkuMapping(
                trading_partner_id=partner.id,
                buyer_sku=row["buyer_sku"],
                buyer_sku_description=row["buyer_sku_description"],
                material_id=material.id if material else None,
                qty_per_buyer_uom=1,
                buyer_uom="EA",
                mapping_status=status,
                confidence_score=1.0 if material else None,
                notes=notes,
            ))
            created += 1
        else:
            sm.buyer_sku_description = row["buyer_sku_description"]
            sm.material_id = material.id if material else None
            sm.mapping_status = status
            sm.confidence_score = 1.0 if material else None
            sm.notes = notes
            sm.deleted_at = None
            updated += 1
    session.flush()
    return created, updated, unmapped


def main() -> None:
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    mapping_path, master_path = Path(sys.argv[1]), Path(sys.argv[2])

    items = _load_sku_master(master_path)
    mappings = _load_mapping(mapping_path)
    print(f"Read {len(items)} materials, {len(mappings)} platform mappings")

    with SyncSessionLocal() as session:
        m_created, m_updated, m_removed = import_materials(session, items)
        s_created, s_updated, s_unmapped = import_mappings(session, mappings)
        session.commit()

    print(f"MaterialMaster: {m_created} created, {m_updated} updated, {m_removed} soft-deleted")
    print(f"SkuMapping:     {s_created} created, {s_updated} updated, {s_unmapped} left UNMAPPED (no SAP code)")


if __name__ == "__main__":
    main()
