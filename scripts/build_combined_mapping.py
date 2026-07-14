"""
Build one combined master-data file from Mapping.xlsx + sku master.xlsx.

Usage:
    python scripts/build_combined_mapping.py <Mapping.xlsx> <sku_master.xlsx> <output.xlsx>

Joins each (platform, platform SKU) mapping row to the SKU master on SAP item
code, pulling in CASE SIZE / EAN / HSN / grammage. Rows whose SAP code is
missing ('#N/A') are highlighted red — those SKUs cannot be validated or
pushed to SAP until ops assigns a SAP code.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

_HEADERS = [
    "CHAIN", "PLATFORM ITEM CODE", "PLATFORM ITEM NAME",
    "SAP ITEM CODE", "SAP NAME", "CASE SIZE", "GRAMMAGE (g)",
    "EAN", "HSN", "MRP", "GST RATE", "DISCOUNT", "UNIT COST", "STATUS",
]
_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def main() -> None:
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)
    mapping_path, master_path, out_path = (Path(p) for p in sys.argv[1:4])

    master_wb = openpyxl.load_workbook(master_path, read_only=True)
    master: dict[str, tuple] = {}
    for r in list(master_wb.active.iter_rows(values_only=True))[1:]:
        if r[3]:
            master[str(r[3]).strip().upper()] = r

    mapping_wb = openpyxl.load_workbook(mapping_path, read_only=True)
    mapping_rows = [r for r in list(mapping_wb.active.iter_rows(values_only=True))[1:] if r[0]]

    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "Combined Mapping"
    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    n_missing_sap = n_missing_master = 0
    for r in mapping_rows:
        sap_code = str(r[4]).strip().upper() if r[4] else ""
        m = master.get(sap_code)
        if sap_code in ("", "#N/A", "N/A", "#REF!"):
            status, fill = "MISSING SAP CODE — assign in sku master", _RED
            sap_code, m = "", None
            n_missing_sap += 1
        elif m is None:
            status, fill = "SAP CODE NOT IN SKU MASTER", _YELLOW
            n_missing_master += 1
        else:
            status, fill = "OK", None

        ws.append([
            r[0],                       # CHAIN
            r[1],                       # PLATFORM ITEM CODE
            r[2],                       # PLATFORM ITEM NAME
            sap_code or None,           # SAP ITEM CODE
            (m[2] if m else r[3]),      # SAP NAME (prefer sku master)
            (m[5] if m else None),      # CASE SIZE
            (m[4] if m else None),      # GRAMMAGE
            (m[7] if m else None),      # EAN
            (m[8] if m else None),      # HSN
            (m[6] if m else r[6]),      # MRP (prefer sku master)
            r[8],                       # GST RATE
            r[7],                       # DISCOUNT
            r[9],                       # UNIT COST
            status,
        ])
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    out.save(out_path)
    print(f"Wrote {ws.max_row - 1} rows to {out_path}")
    print(f"  missing SAP code (red):        {n_missing_sap}")
    print(f"  SAP code not in master (yellow): {n_missing_master}")


if __name__ == "__main__":
    main()
